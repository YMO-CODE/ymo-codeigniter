#!/usr/bin/env python3
"""Merge YMO Customer data sheet.xlsx tabs into one deduped contacts CSV for CRM import.

Usage:
  python import/scripts/merge_ymo_customer_excel.py "path/to/YMO Customer data sheet.xlsx"
  python import/scripts/merge_ymo_customer_excel.py   # default: import/source/YMO Customer data sheet.xlsx

Output:
  import/staging/contacts_master.csv
  import/staging/contacts_merge_report.json
"""
from __future__ import annotations

import csv
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = ROOT / "import" / "source" / "YMO Customer data sheet.xlsx"
OUT_CSV = ROOT / "import" / "staging" / "contacts_master.csv"
OUT_REPORT = ROOT / "import" / "staging" / "contacts_merge_report.json"

SHEET_PUNE = {"G1 Pune old", "G2 Pune old "}
SHEET_FEB = "Feb21 - Oct 22 old "
SHEET_WAKAD = "wakad exhibition list 2026 MAY"

TAG_BY_SHEET = {
    "G1 Pune old": "pune-g1",
    "G2 Pune old ": "pune-g2",
    SHEET_FEB: "feb21-oct22",
    SHEET_WAKAD: "exhibition-wakad-2026",
}


def norm_mobile(raw) -> str:
    if raw is None:
        return ""
    s = re.sub(r"\D", "", str(raw).strip())
    if s.lower() in ("na", "n/a", "none", "nil", "-"):
        return ""
    if s.startswith("91") and len(s) == 12:
        s = s[2:]
    if len(s) == 11 and s.startswith("0"):
        s = s[1:]
    if len(s) == 10:
        return s
    return ""


def norm_email(raw) -> str:
    if raw is None:
        return ""
    s = str(raw).strip().lower()
    if s in ("", "na", "n/a", "none"):
        return ""
    if "@" not in s or " " in s:
        return ""
    return s[:180]


def norm_name(raw) -> str:
    if raw is None:
        return ""
    s = re.sub(r"\s+", " ", str(raw).strip())
    if s.lower() in ("customer name", "name", "na", "n/a"):
        return ""
    return s[:120]


def fmt_date(raw) -> str:
    if raw is None or str(raw).strip() == "":
        return ""
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    return str(raw).strip()[:32]


def cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def visit_key(sheet: str, visit_line: str) -> str:
    return f"{sheet}|{visit_line}"


def parse_pune_row(sheet: str, row) -> dict | None:
    name = norm_name(cell(row, 3))
    mobile = norm_mobile(cell(row, 4))
    email = norm_email(cell(row, 5))
    if not name and not mobile:
        return None
    parts = []
    d = fmt_date(cell(row, 1))
    vehicle = str(cell(row, 6) or "").strip()
    veh_no = str(cell(row, 7) or "").strip()
    km = str(cell(row, 8) or "").strip()
    work_need = str(cell(row, 9) or "").strip()
    work_done = str(cell(row, 11) or "").strip()
    pending = str(cell(row, 13) or "").strip()
    assigned = str(cell(row, 14) or "").strip()
    if vehicle or veh_no:
        parts.append(f"Vehicle: {vehicle} {veh_no}".strip())
    if km and km not in ("na", "NA", "-"):
        parts.append(f"km: {km}")
    if work_done:
        parts.append(f"Done: {work_done}")
    elif work_need:
        parts.append(f"Requested: {work_need}")
    if pending and pending.lower() not in ("na", "n/a", "-", ""):
        parts.append(f"Pending: {pending}")
    if assigned:
        parts.append(f"Assigned: {assigned}")
    line = " | ".join(p for p in parts if p)
    tag = TAG_BY_SHEET.get(sheet, sheet)
    visit = f"[{d} {tag}] {line}".strip() if d else f"[{tag}] {line}".strip()
    return {"name": name, "mobile": mobile, "email": email, "visit": visit, "tag": tag}


def parse_feb_row(row) -> dict | None:
    name = norm_name(cell(row, 1))
    mobile = norm_mobile(cell(row, 2))
    if not name and not mobile:
        return None
    garage = str(cell(row, 3) or "").strip()
    vehicle = str(cell(row, 4) or "").strip()
    veh_no = str(cell(row, 5) or "").strip()
    work = str(cell(row, 6) or "").strip()
    d = fmt_date(cell(row, 0))
    parts = []
    if garage:
        parts.append(f"Garage: {garage}")
    if vehicle or veh_no:
        parts.append(f"Vehicle: {vehicle} {veh_no}".strip())
    if work:
        parts.append(f"Done: {work}")
    line = " | ".join(p for p in parts if p)
    tag = TAG_BY_SHEET[SHEET_FEB]
    if garage.upper() in ("G1", "G2"):
        tag = f"pune-{garage.lower()}"
    visit = f"[{d or tag} {tag}] {line}".strip()
    return {"name": name, "mobile": mobile, "email": "", "visit": visit, "tag": tag}


def parse_wakad_row(row) -> dict | None:
    name = norm_name(cell(row, 0))
    mobile = norm_mobile(cell(row, 2))
    car = str(cell(row, 1) or "").strip()
    if not name and not mobile:
        return None
    tag = TAG_BY_SHEET[SHEET_WAKAD]
    visit = f"[{tag}] Vehicle interest: {car}".strip() if car else f"[{tag}] Exhibition lead"
    return {"name": name, "mobile": mobile, "email": "", "visit": visit, "tag": tag}


def merge_record(store: dict, key: str, rec: dict, stats: dict):
    if key not in store:
        store[key] = {
            "name": rec["name"] or "Unknown",
            "mobile": rec["mobile"],
            "email": rec["email"],
            "notes_lines": [],
            "visit_keys": set(),
            "tags": set(),
        }
        stats["unique_keys"] += 1
    entry = store[key]
    if rec["name"] and (entry["name"] == "Unknown" or len(rec["name"]) > len(entry["name"])):
        entry["name"] = rec["name"]
    if rec["email"] and not entry["email"]:
        entry["email"] = rec["email"]
    if rec["mobile"] and not entry["mobile"]:
        entry["mobile"] = rec["mobile"]
    entry["tags"].add(rec["tag"])
    vk = visit_key(rec["tag"], rec["visit"])
    if vk not in entry["visit_keys"]:
        entry["visit_keys"].add(vk)
        entry["notes_lines"].append(rec["visit"])
        stats["visits_appended"] += 1
    else:
        stats["visits_deduped"] += 1


def raw_phone_invalid(raw, normalized: str) -> bool:
    if normalized:
        return False
    if raw is None:
        return False
    s = str(raw).strip()
    if s == "" or s.lower() in ("na", "n/a", "none", "nil", "-"):
        return False
    return True


def note_invalid_phone(stats: dict, raw, name: str, sheet: str):
    stats["invalid_phones"] += 1
    if len(stats["invalid_phone_samples"]) < 20:
        stats["invalid_phone_samples"].append({
            "sheet": sheet,
            "name": name,
            "raw_phone": str(raw).strip()[:32],
        })


def store_key(rec: dict) -> str | None:
    if rec["mobile"]:
        return f"m:{rec['mobile']}"
    name = rec["name"].lower().strip()
    if name:
        return f"n:{name}"
    return None


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        print(f"File not found: {xlsx}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    store: dict[str, dict] = {}
    stats = {
        "rows_read": 0,
        "rows_skipped": 0,
        "unique_keys": 0,
        "visits_appended": 0,
        "visits_deduped": 0,
        "no_key_skipped": 0,
        "invalid_phones": 0,
        "name_only_contacts": 0,
        "invalid_phone_samples": [],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name in ("LR exhibation feb 26", "Sheet3"):
            continue
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if sheet_name in SHEET_PUNE:
                if i == 0:
                    continue
                raw_phone = cell(row, 4)
                rec = parse_pune_row(sheet_name, row)
            elif sheet_name == SHEET_FEB:
                if i == 0:
                    continue
                raw_phone = cell(row, 2)
                rec = parse_feb_row(row)
            elif sheet_name == SHEET_WAKAD:
                if i == 0:
                    continue
                raw_phone = cell(row, 2)
                rec = parse_wakad_row(row)
            else:
                continue
            stats["rows_read"] += 1
            if not rec:
                stats["rows_skipped"] += 1
                continue
            if raw_phone_invalid(raw_phone, rec["mobile"]):
                note_invalid_phone(stats, raw_phone, rec["name"], sheet_name)
            key = store_key(rec)
            if not key:
                stats["no_key_skipped"] += 1
                continue
            merge_record(store, key, rec, stats)

    wb.close()

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    rows_out = []
    for entry in store.values():
        notes = "\n".join(entry["notes_lines"])
        tags = ",".join(sorted(entry["tags"]))
        if not entry["mobile"]:
            stats["name_only_contacts"] += 1
        rows_out.append({
            "name": entry["name"],
            "mobile": entry["mobile"],
            "email": entry["email"],
            "workshop": "",
            "notes": notes,
            "tags": tags,
        })
    rows_out.sort(key=lambda r: (r["name"].lower(), r["mobile"]))

    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["name", "mobile", "email", "workshop", "notes", "tags"])
        w.writeheader()
        w.writerows(rows_out)

    report = {
        "source": str(xlsx),
        "output_csv": str(OUT_CSV),
        "contacts_written": len(rows_out),
        **stats,
    }
    OUT_REPORT.write_text(json.dumps(report, indent=2), encoding="utf-8")

    print(json.dumps(report, indent=2))
    print(f"\nWrote {OUT_CSV} ({len(rows_out)} contacts)")


if __name__ == "__main__":
    main()
